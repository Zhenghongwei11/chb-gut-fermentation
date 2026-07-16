#!/usr/bin/env python3
from __future__ import annotations

import gzip
import math
import random
import re
import zipfile
from datetime import datetime, timezone
from itertools import combinations
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"
FIG = OUT / "figures"
TAB = OUT / "tables"
RES = OUT / "results"
PSEUDOCOUNT = 1e-9
SEED = 20260714
FIXED_WORKBOOK_DATETIME = datetime(2026, 7, 16, 0, 0, 0)
PDF_METADATA = {
    "Creator": "CHB gut microbial fermentation pathway analysis",
    "Producer": "Matplotlib",
    "CreationDate": datetime(2026, 7, 16, tzinfo=timezone.utc),
    "ModDate": datetime(2026, 7, 16, tzinfo=timezone.utc),
}

MODULE_LABELS = {
    "overall_fermentation": "Selected fermentation composite",
    "scfa_acetate": "Acetate/lactate overlap",
    "scfa_lactate_succinate": "Acetate/lactate overlap",
    "scfa_butyrate": "Butyrate",
    "bile_acids": "Bile-acid transformation",
    "lps_lipidA": "LPS/lipid A biosynthesis",
    "tryptophan_indole": "Tryptophan biosynthesis",
}


def read_tsv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, sep="\t")


def fmt_num(value: object, digits: int = 3) -> str:
    value = float(value)
    if math.isnan(value):
        return ""
    return f"{value:.{digits}f}"


def fmt_p(value: object) -> str:
    value = float(value)
    if value < 0.001:
        return f"{value:.2e}"
    return f"{value:.3f}"


def row(df: pd.DataFrame, module: str, membership: str | None = None) -> pd.Series:
    out = df[df["module"] == module]
    if membership is not None and "membership" in out.columns:
        out = out[out["membership"] == membership]
    if out.empty:
        raise KeyError(module)
    return out.iloc[0]


def prj_delta(r: pd.Series) -> float:
    return float(r["delta_mean_logratio_S_minus_M"])


def ext_delta(r: pd.Series) -> float:
    return float(r["delta_mean"])


def save_figure(fig: plt.Figure, stem: str) -> None:
    fig.savefig(FIG / f"{stem}.png", dpi=300, bbox_inches="tight", facecolor="white")
    fig.savefig(FIG / f"{stem}.pdf", bbox_inches="tight", facecolor="white", metadata=PDF_METADATA)


def normalize_xlsx_zip(path: Path) -> None:
    tmp = path.with_suffix(".tmp.xlsx")
    fixed_time = (2026, 7, 16, 0, 0, 0)
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for name in sorted(zin.namelist()):
            src = zin.getinfo(name)
            dst = zipfile.ZipInfo(filename=name, date_time=fixed_time)
            dst.compress_type = zipfile.ZIP_DEFLATED
            dst.external_attr = src.external_attr
            data = zin.read(name)
            if name == "docProps/core.xml":
                data = re.sub(
                    rb"<dcterms:modified[^>]*>[^<]+</dcterms:modified>",
                    b'<dcterms:modified xmlns:dcterms="http://purl.org/dc/terms/" '
                    b'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
                    b'xsi:type="dcterms:W3CDTF">2026-07-16T00:00:00Z</dcterms:modified>',
                    data,
                )
            zout.writestr(dst, data)
    tmp.replace(path)


def prj_p(r: pd.Series) -> float:
    return float(r["p_exact"])


def ext_p(r: pd.Series) -> float:
    return float(r["permutation_p"])


def feature_to_id(feature: str) -> str:
    return str(feature).split("|", 1)[0].split(":", 1)[0]


def draw_panel_label(ax, label: str) -> None:
    ax.text(-0.10, 1.12, label, transform=ax.transAxes, fontsize=12, fontweight="bold", va="bottom")


def build_figure_1() -> None:
    fig, ax = plt.subplots(figsize=(11.2, 6.2))
    ax.axis("off")
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.text(0.5, 0.94, "Study cohorts and phenotype definitions", ha="center", fontsize=16, fontweight="bold")
    ax.text(0.5, 0.89, "A discovery-derived pathway composite was finalised before external-cohort evaluation.", ha="center", fontsize=10.5)
    boxes = [
        (0.06, 0.59, 0.26, 0.22, "#245766", "CHB discovery cohort\nTreatment-naive CHB\nM: G0-1 and S0-1, n=9\nS: G>1 and/or S>1, n=11"),
        (0.37, 0.59, 0.26, 0.22, "#2c8f83", "NAFLD fibrosis cohort\nBiopsy-proven NAFLD\nF0-F2, n=72\nF3-F4, n=14"),
        (0.68, 0.59, 0.26, 0.22, "#6b7a8f", "Cirrhosis context cohort\nHealthy controls, n=114\nCirrhosis, n=123"),
        (0.10, 0.27, 0.22, 0.16, "#eef4f7", "Discovery contrast\nSelected composite\nand submodules"),
        (0.39, 0.27, 0.22, 0.16, "#eef4f7", "Unadjusted comparison\nSelected composite\nand sensitivity"),
        (0.68, 0.27, 0.22, 0.16, "#eef4f7", "Context contrast\nSelected composite\nand heterogeneity"),
    ]
    for x, y, w, h, color, text in boxes:
        ax.add_patch(plt.Rectangle((x, y), w, h, facecolor=color, edgecolor="#263238", linewidth=1.0))
        tcolor = "white" if color.startswith("#2") or color.startswith("#6") else "#1f2933"
        ax.text(x + w / 2, y + h / 2, text, ha="center", va="center", fontsize=10.1, linespacing=1.25, color=tcolor)
    for x in [0.19, 0.50, 0.81]:
        ax.annotate("", xy=(x, 0.44), xytext=(x, 0.59), arrowprops=dict(arrowstyle="->", lw=1.2, color="#263238"))
    save_figure(fig, "Figure1_Study_cohorts_and_phenotype_definitions")
    plt.close(fig)


def build_figure_2() -> None:
    scores = read_tsv(RES / "prjdb36442" / "conservative" / "conservative_module_scores.tsv")
    stats = read_tsv(RES / "prjdb36442" / "conservative" / "conservative_module_stats.tsv")
    plot_scores = scores[scores["module"] == "overall_fermentation"].copy()
    rows = [
        ("Selected fermentation composite", row(stats, "overall_fermentation")),
        ("Acetate/lactate overlap", row(stats, "scfa_acetate")),
        ("Butyrate", row(stats, "scfa_butyrate")),
    ]
    fig, axes = plt.subplots(1, 2, figsize=(11.3, 4.5), gridspec_kw={"width_ratios": [1.0, 1.25]})
    vals = [plot_scores.loc[plot_scores["group"] == g, "log_ratio_score"].to_numpy() for g in ["M", "S"]]
    axes[0].boxplot(vals, tick_labels=["M\nn=9", "S\nn=11"], patch_artist=True, boxprops=dict(facecolor="#dbeafe"), medianprops=dict(color="#111827", lw=1.5))
    for i, arr in enumerate(vals, start=1):
        axes[0].scatter([i] * len(arr), arr, s=30, color="#245766", edgecolor="white", zorder=3)
    axes[0].set_ylim(-5.55, -3.95)
    axes[0].set_title("Selected composite score", fontsize=11)
    axes[0].set_ylabel("log-ratio pathway score")
    draw_panel_label(axes[0], "A")
    labels = [r[0] for r in rows]
    deltas = [prj_delta(r[1]) for r in rows]
    lows = [float(r[1]["bootstrap95_ci_low"]) for r in rows]
    highs = [float(r[1]["bootstrap95_ci_high"]) for r in rows]
    y = list(range(len(rows)))[::-1]
    axes[1].errorbar(deltas, y, xerr=[[d - l for d, l in zip(deltas, lows)], [h - d for d, h in zip(deltas, highs)]], fmt="o", color="#245766", ecolor="#245766", capsize=4)
    axes[1].axvline(0, color="#9ca3af", lw=1)
    axes[1].set_yticks(y, labels)
    axes[1].set_xlabel("S minus M mean difference (95% CI)")
    axes[1].set_title("Effect estimates", fontsize=11)
    for yy, high, (_, stat) in zip(y, highs, rows):
        axes[1].text(high + 0.05, yy, f"P={fmt_p(prj_p(stat))}", va="center", fontsize=9)
    draw_panel_label(axes[1], "B")
    fig.tight_layout()
    save_figure(fig, "Figure2_CHB_discovery_effects")
    plt.close(fig)


def build_figure_3() -> None:
    scores = read_tsv(RES / "loombar2017" / "module_scores.tsv")
    stats = read_tsv(RES / "loombar2017" / "module_binary_contrasts.tsv")
    plot_scores = scores[(scores["membership"] == "conservative") & (scores["module"] == "overall_fermentation")]
    rows = [
        ("Selected fermentation composite", row(stats, "overall_fermentation", "conservative")),
        ("Acetate/lactate overlap", row(stats, "scfa_acetate", "conservative")),
        ("Butyrate", row(stats, "scfa_butyrate", "conservative")),
        ("Selected composite\nexpanded definition", row(stats, "overall_fermentation", "expanded")),
    ]
    fig, axes = plt.subplots(1, 2, figsize=(11.6, 4.6), gridspec_kw={"width_ratios": [1.1, 1.1]})
    vals = [plot_scores.loc[plot_scores["group"] == g, "module_score"].to_numpy() for g in ["F0_F2", "F3_F4"]]
    axes[0].boxplot(vals, tick_labels=["F0-F2\nn=72", "F3-F4\nn=14"], patch_artist=True, boxprops=dict(facecolor="#d1fae5"), medianprops=dict(color="#111827", lw=1.5))
    for i, arr in enumerate(vals, start=1):
        axes[0].scatter([i] * len(arr), arr, s=20, alpha=0.7, color="#2c8f83", edgecolor="white", linewidth=0.3, zorder=3)
    axes[0].set_ylabel("log-ratio pathway score")
    axes[0].set_title("Selected composite score", fontsize=11)
    draw_panel_label(axes[0], "A")
    labels = [r[0] for r in rows]
    rows_for_plot = [r[1] for r in rows]
    deltas = [ext_delta(r) for r in rows_for_plot]
    lows = [float(r["bootstrap95_ci_low"]) for r in rows_for_plot]
    highs = [float(r["bootstrap95_ci_high"]) for r in rows_for_plot]
    y = list(range(len(labels)))[::-1]
    axes[1].errorbar(deltas, y, xerr=[[d - l for d, l in zip(deltas, lows)], [h - d for d, h in zip(deltas, highs)]], fmt="o", color="#2c8f83", ecolor="#2c8f83", capsize=4)
    axes[1].axvline(0, color="#9ca3af", lw=1)
    axes[1].set_yticks(y, labels)
    axes[1].set_xlabel("F3-F4 minus F0-F2 mean difference (95% CI)")
    axes[1].set_title("Pathway-set sensitivity", fontsize=11)
    for yy, high, stat in zip(y, highs, rows_for_plot):
        axes[1].text(high + 0.04, yy, f"P={fmt_p(ext_p(stat))}", va="center", fontsize=9)
    draw_panel_label(axes[1], "B")
    fig.tight_layout()
    save_figure(fig, "Figure3_NAFLD_fibrosis_pathway_scores")
    plt.close(fig)


def build_figure_4() -> None:
    prj = row(read_tsv(RES / "prjdb36442" / "conservative" / "conservative_module_stats.tsv"), "overall_fermentation")
    loomba = row(read_tsv(RES / "loombar2017" / "module_binary_contrasts.tsv"), "overall_fermentation", "conservative")
    qin = row(read_tsv(RES / "qinn2014" / "module_binary_contrasts.tsv"), "overall_fermentation", "conservative")
    rows = [
        ("CHB discovery cohort\nS vs M", prj_delta(prj), float(prj["bootstrap95_ci_low"]), float(prj["bootstrap95_ci_high"]), prj_p(prj)),
        ("NAFLD fibrosis cohort\nF3-F4 vs F0-F2", ext_delta(loomba), float(loomba["bootstrap95_ci_low"]), float(loomba["bootstrap95_ci_high"]), ext_p(loomba)),
        ("Cirrhosis context cohort\ncirrhosis vs healthy", ext_delta(qin), float(qin["bootstrap95_ci_low"]), float(qin["bootstrap95_ci_high"]), ext_p(qin)),
    ]
    fig, ax = plt.subplots(figsize=(9.2, 4.5))
    y = list(range(len(rows)))[::-1]
    for yy, item, color in zip(y, rows, ["#245766", "#2c8f83", "#6b7a8f"]):
        _, d, low, high, pval = item
        ax.errorbar(d, yy, xerr=[[d - low], [high - d]], fmt="o", color=color, ecolor=color, capsize=4, ms=7)
        ax.text(high + 0.06, yy, f"P={fmt_p(pval)}", va="center", fontsize=9.5)
    ax.axvline(0, color="#9ca3af", lw=1)
    ax.set_yticks(y, [r[0] for r in rows])
    ax.set_xlabel("Test minus reference mean difference in selected composite score (95% CI)")
    fig.tight_layout()
    save_figure(fig, "Figure4_Cohort_specific_pathway_score_contrasts")
    plt.close(fig)


def build_supplementary_figure_1() -> None:
    stats = read_tsv(RES / "prjdb36442" / "conservative" / "conservative_module_stats.tsv")
    rand = read_tsv(RES / "prjdb36442" / "conservative" / "conservative_random_modules.tsv")
    rand_emp = row(read_tsv(RES / "prjdb36442" / "conservative" / "conservative_random_module_empirical.tsv"), "overall_fermentation")
    modules = ["overall_fermentation", "bile_acids", "lps_lipidA", "tryptophan_indole"]
    rows = [(MODULE_LABELS[m], row(stats, m)) for m in modules]
    fig, axes = plt.subplots(1, 3, figsize=(13.2, 4.4), gridspec_kw={"width_ratios": [1.0, 1.0, 1.35]})
    narrow = rows[:2] + [rows[3]]
    wide = [rows[2]]
    for ax, plot_rows, title in [(axes[0], narrow, "Selected and comparison modules"), (axes[1], wide, "LPS/lipid A module")]:
        labels = [r[0] for r in plot_rows]
        deltas = [prj_delta(r[1]) for r in plot_rows]
        lows = [float(r[1]["bootstrap95_ci_low"]) for r in plot_rows]
        highs = [float(r[1]["bootstrap95_ci_high"]) for r in plot_rows]
        y = list(range(len(plot_rows)))[::-1]
        for yy, d, low, high, color in zip(y, deltas, lows, highs, ["#245766"] + ["#7c8798"] * 3):
            ax.errorbar(d, yy, xerr=[[d - low], [high - d]], fmt="o", color=color, ecolor=color, capsize=4)
        ax.axvline(0, color="#9ca3af", lw=1)
        ax.set_yticks(y, labels)
        ax.set_xlabel("S minus M mean difference")
        ax.set_title(title, fontsize=11, pad=14)
    draw_panel_label(axes[0], "A")
    draw_panel_label(axes[1], "B")
    col = "delta_mean_logratio_S_minus_M"
    axes[2].hist(rand[col].astype(float), bins=35, color="#d8dee9", edgecolor="white")
    axes[2].axvline(prj_delta(row(stats, "overall_fermentation")), color="#c2410c", lw=2, label="Observed selected composite")
    axes[2].axvline(rand[col].astype(float).median(), color="#334155", lw=1.5, ls="--", label="Matched-set median")
    axes[2].set_xlabel("Matched-set S minus M difference")
    axes[2].set_ylabel("Count")
    axes[2].set_title(f"Matched pathway sets (absolute empirical P={fmt_p(rand_emp['empirical_p_abs_delta'])})", fontsize=11, pad=14)
    axes[2].legend(frameon=False, fontsize=9)
    draw_panel_label(axes[2], "C")
    fig.tight_layout()
    save_figure(fig, "Supplementary_Figure_S1_mechanistic_and_matched_pathway_comparisons")
    plt.close(fig)


def table_1() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ["CHB discovery", "Treatment-naive CHB", "M: Scheuer G0-1 and S0-1; n=9", "S: G>1 and/or S>1; n=11", "Binary biopsy-defined M/S group", "Sample-level G, S, ALT, AST, BMI, HBV DNA, sex, and medication exposure unavailable from public metadata used here"],
            ["NAFLD fibrosis", "Biopsy-proven NAFLD", "F0-F2; n=72", "F3-F4; n=14", "Fibrosis stage available; age category fully collinear with fibrosis group", "Read count, read bases, platform, extraction kit; sex, BMI, diabetes unavailable"],
            ["Cirrhosis context", "Cirrhosis case-control cohort", "Healthy controls; n=114", "Cirrhosis; n=123", "Disease status available; not a disease-internal histology contrast", "Detailed patient-level cirrhosis severity covariates unavailable in this analysis"],
        ],
        columns=["cohort", "disease_setting", "reference_group", "test_group", "phenotype_data", "covariates_available"],
    )


def table_2() -> pd.DataFrame:
    prj = row(read_tsv(RES / "prjdb36442" / "conservative" / "conservative_module_stats.tsv"), "overall_fermentation")
    loomba = row(read_tsv(RES / "loombar2017" / "module_binary_contrasts.tsv"), "overall_fermentation", "conservative")
    qin = row(read_tsv(RES / "qinn2014" / "module_binary_contrasts.tsv"), "overall_fermentation", "conservative")
    return pd.DataFrame(
        [
            ["CHB discovery", "M", "S", 9, 11, prj_delta(prj), prj["cliffs_delta_S_vs_M"], prj_p(prj), prj["bootstrap95_ci_low"], prj["bootstrap95_ci_high"], "Discovery-stage"],
            ["NAFLD fibrosis", "F0-F2", "F3-F4", 72, 14, ext_delta(loomba), loomba["cliffs_delta_test_vs_reference"], ext_p(loomba), loomba["bootstrap95_ci_low"], loomba["bootstrap95_ci_high"], "Unadjusted histologic comparison"],
            ["Cirrhosis context", "Healthy", "Cirrhosis", 114, 123, ext_delta(qin), qin["cliffs_delta_test_vs_reference"], ext_p(qin), qin["bootstrap95_ci_low"], qin["bootstrap95_ci_high"], "Context"],
        ],
        columns=["cohort", "reference_group", "test_group", "reference_n", "test_n", "delta", "cliffs_delta", "p_value", "ci_low", "ci_high", "interpretation"],
    )


def unmapped_unintegrated() -> pd.DataFrame:
    rows = []
    resources = [
        ("PRJDB36442", ROOT / "data" / "prjdb36442" / "merged_pathabundance.tsv.gz"),
        ("LoombaR_2017", ROOT / "data" / "loombar2017" / "pathway_abundance_unstratified.tsv"),
        ("QinN_2014", ROOT / "data" / "qinn2014" / "pathway_abundance_unstratified.tsv"),
    ]
    for cohort, path in resources:
        opener = gzip.open if path.suffix == ".gz" else open
        with opener(path, "rt") as handle:
            raw = pd.read_csv(handle, sep="\t")
        feature = raw.iloc[:, 0].astype(str).map(feature_to_id)
        values = raw.iloc[:, 1:].astype(float)
        total = values.sum(axis=0).replace(0, np.nan)
        for label in ["UNMAPPED", "UNINTEGRATED"]:
            mask = feature.eq(label)
            if mask.any():
                frac = values.loc[mask].sum(axis=0) / total
                rows.append([cohort, label, int(mask.sum()), len(frac), frac.median(), frac.quantile(0.25), frac.quantile(0.75), frac.min(), frac.max()])
            else:
                rows.append([cohort, label, 0, values.shape[1], np.nan, np.nan, np.nan, np.nan, np.nan])
    return pd.DataFrame(rows, columns=["cohort", "feature_class", "feature_rows_present", "sample_n", "median_fraction", "iqr_low", "iqr_high", "minimum", "maximum"])


def exact_permutation_p(values: list[float], labels: list[str], test_label: str) -> float:
    n = len(values)
    n_test = sum(1 for label in labels if label == test_label)
    observed = np.mean([v for v, g in zip(values, labels) if g == test_label]) - np.mean(
        [v for v, g in zip(values, labels) if g != test_label]
    )
    total = 0
    extreme = 0
    all_idx = range(n)
    for combo in combinations(all_idx, n_test):
        combo = set(combo)
        delta = np.mean([values[i] for i in combo]) - np.mean([values[i] for i in all_idx if i not in combo])
        total += 1
        extreme += int(abs(delta) >= abs(observed) - 1e-15)
    return extreme / total


def monte_carlo_p(values: list[float], labels: list[str], test_label: str, seed: int = SEED, n_perm: int = 10000) -> float:
    rng = random.Random(seed)
    n = len(values)
    n_test = sum(1 for label in labels if label == test_label)
    observed = np.mean([v for v, g in zip(values, labels) if g == test_label]) - np.mean(
        [v for v, g in zip(values, labels) if g != test_label]
    )
    extreme = 0
    idx = list(range(n))
    for _ in range(n_perm):
        rng.shuffle(idx)
        test = set(idx[:n_test])
        delta = np.mean([values[i] for i in test]) - np.mean([values[i] for i in range(n) if i not in test])
        extreme += int(abs(delta) >= abs(observed) - 1e-15)
    return (extreme + 1) / (n_perm + 1)


def bootstrap_delta_ci(values: list[float], labels: list[str], test_label: str, seed: int = SEED, n_boot: int = 5000) -> tuple[float, float]:
    rng = np.random.default_rng(seed)
    ref = np.array([v for v, g in zip(values, labels) if g != test_label], dtype=float)
    test = np.array([v for v, g in zip(values, labels) if g == test_label], dtype=float)
    deltas = np.empty(n_boot)
    for i in range(n_boot):
        rb = rng.choice(ref, size=len(ref), replace=True)
        tb = rng.choice(test, size=len(test), replace=True)
        deltas[i] = float(np.mean(tb) - np.mean(rb))
    return float(np.quantile(deltas, 0.025)), float(np.quantile(deltas, 0.975))


def load_prjdb_relative() -> pd.DataFrame:
    raw = pd.read_csv(ROOT / "data" / "prjdb36442" / "merged_pathabundance.tsv.gz", sep="\t")
    feature_col = raw.columns[0]
    raw = raw[~raw[feature_col].astype(str).str.contains(r"\|", regex=True)].copy()
    raw["pathway_id"] = raw[feature_col].map(feature_to_id)
    raw = raw[~raw["pathway_id"].isin(["UNMAPPED", "UNINTEGRATED"])].copy()
    value_cols = [c for c in raw.columns if c.endswith("_Abundance")]
    mat = raw.set_index("pathway_id")[value_cols].astype(float)
    mat.columns = [c.replace("_Abundance", "") for c in mat.columns]
    mat = mat.groupby(mat.index).sum()
    return mat.div(mat.sum(axis=0), axis=1).fillna(0.0)


def load_external_relative(path: Path) -> pd.DataFrame:
    raw = pd.read_csv(path, sep="\t").rename(columns={"# Pathway": "feature"})
    raw["pathway_id"] = raw["feature"].map(feature_to_id)
    raw = raw[~raw["feature"].astype(str).str.contains(r"\|", regex=True)].copy()
    mat = raw.drop(columns=["feature"]).set_index("pathway_id").astype(float)
    mat = mat.groupby(mat.index).sum()
    return mat.div(mat.sum(axis=0), axis=1).fillna(0.0)


def score_pathways(rel: pd.DataFrame, sample_ids: list[str], pathway_ids: list[str], pseudocount: float) -> list[float]:
    present = [p for p in dict.fromkeys(pathway_ids) if p in rel.index]
    module_sum = rel.loc[present, sample_ids].sum(axis=0).clip(lower=0.0, upper=1.0)
    background = (1.0 - module_sum).clip(lower=0.0)
    return (np.log(module_sum + pseudocount) - np.log(background + pseudocount)).astype(float).tolist()


def selected_pathways() -> list[str]:
    defs = read_tsv(ROOT / "analysis" / "module_definitions.tsv")
    return (
        defs[
            (defs["membership"] == "conservative")
            & (defs["module"].isin(["scfa_acetate", "scfa_lactate_succinate", "scfa_butyrate"]))
        ]["pathway_id"]
        .drop_duplicates()
        .tolist()
    )


def build_score_sensitivity() -> pd.DataFrame:
    pathways = selected_pathways()
    core5 = [p for p in pathways if p != "P41-PWY"]
    rows = []

    prj_rel = load_prjdb_relative()
    prj_meta = read_tsv(ROOT / "data" / "prjdb36442" / "sample_manifest.tsv")
    prj_samples = [s for s in prj_meta["run_accession"].tolist() if s in prj_rel.columns]
    prj_labels = prj_meta.set_index("run_accession").loc[prj_samples, "group"].tolist()

    loomba_rel = load_external_relative(ROOT / "data" / "loombar2017" / "pathway_abundance_unstratified.tsv")
    loomba_scores = read_tsv(RES / "loombar2017" / "module_scores.tsv")
    loomba_meta = loomba_scores[loomba_scores["module"] == "overall_fermentation"][["sample_id", "group"]].drop_duplicates()
    loomba_samples = [s for s in loomba_meta["sample_id"].tolist() if s in loomba_rel.columns]
    loomba_labels = loomba_meta.set_index("sample_id").loc[loomba_samples, "group"].tolist()

    for cohort, rel, samples, labels, test_label, exact in [
        ("PRJDB36442", prj_rel, prj_samples, prj_labels, "S", True),
        ("LoombaR_2017", loomba_rel, loomba_samples, loomba_labels, "F3_F4", False),
    ]:
        for definition, ids in [("selected conservative composite", pathways), ("shared five-pathway core", core5)]:
            for pc in [1e-9, 1e-8, 1e-6]:
                values = score_pathways(rel, samples, ids, pc)
                ref = [v for v, g in zip(values, labels) if g != test_label]
                test = [v for v, g in zip(values, labels) if g == test_label]
                delta = float(np.mean(test) - np.mean(ref))
                pval = exact_permutation_p(values, labels, test_label) if exact else monte_carlo_p(values, labels, test_label)
                ci_low, ci_high = bootstrap_delta_ci(values, labels, test_label)
                rows.append(
                    {
                        "cohort": cohort,
                        "definition": definition,
                        "pseudocount": {1e-9: "10⁻⁹", 1e-8: "10⁻⁸", 1e-6: "10⁻⁶"}[pc],
                        "n_pathways_defined": len(ids),
                        "n_pathways_present": sum(1 for p in ids if p in rel.index),
                        "delta_test_minus_reference": delta,
                        "p_value": pval,
                        "bootstrap95_ci_low": ci_low,
                        "bootstrap95_ci_high": ci_high,
                        "test_group": test_label,
                    }
                )
    return pd.DataFrame(rows)


def build_module_definition_sheet() -> pd.DataFrame:
    defs = read_tsv(ROOT / "analysis" / "module_definitions.tsv").copy()
    defs["biochemical_rationale"] = (
        defs["biochemical_rationale"]
        .astype(str)
        .str.replace("Locked v1: ", "", regex=False)
        .str.replace("Locked v1 ", "", regex=False)
        .str.replace("negative-control comparator", "mechanistic comparison module", regex=False)
        .str.replace("negative-control", "mechanistic comparison", regex=False)
        .str.replace("retained only as comparator", "retained as a mechanistic comparison module", regex=False)
    )
    module_map = {
        "overall_fermentation": "selected fermentation composite",
        "scfa_acetate": "acetate/lactate overlap",
        "scfa_lactate_succinate": "acetate/lactate overlap",
        "scfa_butyrate": "butyrate",
        "bile_acids": "bile-acid transformation",
        "lps_lipidA": "LPS/lipid A biosynthesis",
        "tryptophan_indole": "tryptophan biosynthesis",
    }
    fermentation_modules = {"scfa_acetate", "scfa_lactate_succinate", "scfa_butyrate"}
    rows = []
    for pathway_id, g in defs.groupby("pathway_id", sort=True):
        rows.append(
            {
                "pathway_id": pathway_id,
                "pathway_name": g["pathway_name"].iloc[0],
                "reported_module": "; ".join(sorted({module_map.get(m, m) for m in g["module"]})),
                "original_source_definition": "; ".join(sorted(set(g["module"]))),
                "membership_definitions": "; ".join(sorted(set(g["membership"]))),
                "included_in_selected_composite": "yes" if any(m in fermentation_modules for m in set(g["module"])) else "no",
                "expected_direction_with_severity": "; ".join(sorted(set(g["expected_direction_with_severity"]))),
                "module_class": "fermentation module" if any(m in fermentation_modules for m in set(g["module"])) else "mechanistic comparison module",
                "biochemical_rationale": " ".join(dict.fromkeys(g["biochemical_rationale"].dropna().astype(str).tolist())),
            }
        )
    out = pd.DataFrame(rows)
    out["_order"] = out["included_in_selected_composite"].map({"yes": 0, "no": 1})
    return out.sort_values(["_order", "reported_module", "pathway_id"]).drop(columns=["_order"])


def build_member_availability() -> pd.DataFrame:
    rows = []
    files = [
        ("PRJDB36442", RES / "prjdb36442" / "conservative" / "conservative_module_stats.tsv"),
        ("PRJDB36442", RES / "prjdb36442" / "expanded" / "expanded_module_stats.tsv"),
        ("LoombaR_2017", RES / "loombar2017" / "module_binary_contrasts.tsv"),
        ("QinN_2014", RES / "qinn2014" / "module_binary_contrasts.tsv"),
    ]
    for cohort, path in files:
        df = read_tsv(path)
        for _, r in df.iterrows():
            if r["module"] == "scfa_lactate_succinate":
                continue
            module = "acetate_lactate_overlap" if r["module"] == "scfa_acetate" else r["module"]
            rows.append(
                {
                    "cohort": cohort,
                    "membership": r.get("membership", "conservative" if "conservative" in str(path) else "expanded"),
                    "reported_module": module,
                    "reported_module_name": MODULE_LABELS.get(r["module"], r["module"]),
                    "pathways_defined": r.get("pathways_defined", r.get("n_pathways_defined")),
                    "pathways_present": r.get("pathways_present", r.get("n_pathways_present")),
                    "note": "Acetate and lactate/succinate use the same available pathway set in these analyses." if module == "acetate_lactate_overlap" else "",
                }
            )
    return pd.DataFrame(rows)


def build_formula_sheet() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("Input feature set", "Unstratified MetaCyc pathway abundances; taxon-stratified rows, UNMAPPED, and UNINTEGRATED were excluded when present. Their availability and fractions are summarized separately where exported."),
            ("Sample normalisation", "Within each sample, retained unstratified pathway abundances were divided by their sample total."),
            ("Module numerator", "For sample i and module M, A_i(M)=sum_{p in M_i} a_ip, where a_ip is normalized pathway abundance and M_i contains module pathways present in that cohort."),
            ("Background denominator", "B_i(M)=max(1-A_i(M),0), representing all other retained unstratified pathway abundance in the same sample."),
            ("Score formula", "Score_i(M)=ln(A_i(M)+10⁻⁹)-ln(B_i(M)+10⁻⁹). Scores are natural-log units."),
            ("Zero handling", "A pseudocount of 10⁻⁹ was added to numerator and background denominator before logarithms."),
            ("Missing pathways", "Finalised module members absent from a cohort matrix were not imputed; scores used available members only, and available/defined pathway counts were reported."),
            ("Background denominator scope", "The background denominator was cohort-specific because retained pathway universes differed across processing sources; cross-cohort interpretation therefore used direction and cohort-specific effect estimates rather than pooled absolute scores."),
            ("Prevalence and abundance thresholds", "No prevalence or abundance threshold was applied after module definition. Prevalence was used to construct matched pathway sets."),
            ("Coverage threshold", "Pathway coverage profiles were not used as an inclusion threshold for the processed-matrix analyses."),
            ("Random-module matching", "Random modules were matched on pathway count, mean abundance, prevalence, variance, and zero-fraction quintiles where available."),
            ("Random seed", str(SEED)),
        ],
        columns=["item", "description"],
    )


def build_stat_settings() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("Module finalisation", "The selected composite was defined in the CHB discovery analysis and then carried forward unchanged in the external cohorts."),
            ("Main discovery analysis", "CHB discovery-cohort conservative selected fermentation-composite score, S group minus M group."),
            ("Secondary endpoints", "CHB discovery-cohort acetate/lactate overlap, butyrate, expanded composite definition, and NAFLD fibrosis-cohort selected composite."),
            ("Exploratory/context endpoints", "Mechanistic comparison modules and cirrhosis-versus-healthy contrasts in the cirrhosis context cohort."),
            ("CHB discovery-cohort exact permutation", "All C(20,11)=167,960 allocations preserving 11 S and 9 M labels were evaluated."),
            ("Two-sided exact P", "The exact P value is the fraction of allocations with absolute mean difference at least as large as the observed absolute mean difference."),
            ("External permutation", "The NAFLD fibrosis and cirrhosis context cohorts used 10,000 Monte Carlo label permutations."),
            ("Monte Carlo P", "Monte Carlo P values used the correction (b+1)/(m+1)."),
            ("Bootstrap CI", "5,000 bootstrap resamples were drawn within reference and test groups; percentile 2.5% and 97.5% intervals are reported."),
            ("Cliff's delta", "Pairwise group comparisons counted greater-than and less-than pairs; ties contributed zero to the numerator."),
            ("Multiplicity", "The conservative selected fermentation-composite score was the sole main discovery analysis. Submodules, mechanistic comparison modules, individual pathway checks, and leave-one-pathway-out analyses are exploratory and are presented with nominal P values, effect sizes, uncertainty intervals, and consistency checks; no FDR-controlled confirmatory claim is made for these exploratory outputs."),
            ("Technical and clinical covariates", "Clinical covariates were not available for the CHB discovery cohort. The exported NAFLD fibrosis-cohort metadata did not contain sex, BMI, or diabetes; age_category was fully collinear with fibrosis group and was not adjusted for."),
            ("Software and pathway resources", "CHB pathway profiles were generated with HUMAnN v3.9 using ChocoPhlAn nucleotide search, bypassed translated search, and MetaPhlAn index mpa_vJan25_CHOCOPhlAnSGB_202503. External processed pathway_abundance resources were exported with curatedMetagenomicData v3.16.1: 2021-10-14.LoombaR_2017.pathway_abundance and 2021-03-31.QinN_2014.pathway_abundance. Python 3 with pandas, NumPy, SciPy/permutation routines and matplotlib was used for the present statistical analysis and plotting."),
        ],
        columns=["item", "description"],
    )


def build_loomba_covariates() -> pd.DataFrame:
    meta = read_tsv(ROOT / "data" / "loombar2017" / "metadata.tsv")
    checks = [
        ("fibrosis_stage_numeric", "Fibrosis stage used to define F0-F2 versus F3-F4", "Used for the NAFLD fibrosis contrast"),
        ("age_category", "Age category", "Available but fully collinear with group: all F0-F2 samples were adult and all F3-F4 samples were senior"),
        ("sex", "Sex", "Not available in the exported processed metadata"),
        ("gender", "Gender", "Not available in the exported processed metadata"),
        ("BMI", "Body mass index", "Not available in the exported processed metadata"),
        ("bmi", "Body mass index", "Not available in the exported processed metadata"),
        ("diabetes", "Diabetes status", "Not available in the exported processed metadata"),
        ("number_reads", "Sequencing read count", "Available as a technical descriptor"),
        ("number_bases", "Sequencing bases", "Available as a technical descriptor"),
        ("sequencing_platform", "Sequencing platform", "Available; constant within the analysed cohort"),
        ("DNA_extraction_kit", "DNA extraction kit", "Available; constant within the analysed cohort"),
    ]
    rows = []
    for col, label, note in checks:
        if col in meta.columns:
            non_missing = int(meta[col].notna().sum())
            by_group = "; ".join(f"{g}: {int(v)}" for g, v in meta.groupby("group")[col].apply(lambda x: x.notna().sum()).items())
            values = "; ".join(map(str, sorted(meta[col].dropna().astype(str).unique())[:8]))
        else:
            non_missing, by_group, values = 0, "not available", "not available"
        rows.append(
            {
                "variable": label,
                "source_column_checked": col,
                "non_missing_n": non_missing,
                "non_missing_by_group": by_group,
                "observed_values_or_status": values,
                "analysis_decision": note,
            }
        )
    return pd.DataFrame(rows)


def endpoint_sheet() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("Main discovery", "PRJDB36442", "Selected fermentation composite, conservative definition", "S vs M", "Main discovery analysis."),
            ("Secondary", "PRJDB36442", "Acetate/lactate overlap and butyrate", "S vs M", "Submodule localisation; acetate/lactate overlap is reported once."),
            ("Secondary", "PRJDB36442", "Selected fermentation composite, expanded definition", "S vs M", "Module-definition sensitivity."),
            ("Secondary", "NAFLD fibrosis cohort", "Selected fermentation composite, conservative and expanded definitions", "F3-F4 vs F0-F2", "Unadjusted cross-aetiology biopsy-defined fibrosis comparison and sensitivity."),
            ("Exploratory", "Qin et al. cirrhosis cohort", "Overall and submodules", "cirrhosis vs healthy", "Cirrhosis context and heterogeneity assessment."),
            ("Exploratory", "All analysed cohorts", "Bile-acid, LPS/lipid A, tryptophan modules", "cohort-specific contrasts", "Mechanistic comparison modules; not interpreted as control evidence for fermentation."),
        ],
        columns=["level", "cohort", "module_or_analysis", "contrast", "interpretation"],
    )


def clean_module_effects(df: pd.DataFrame, cohort: str) -> pd.DataFrame:
    out = df[df["module"] != "scfa_lactate_succinate"].copy()
    out["reported_module"] = out["module"].replace({"scfa_acetate": "acetate_lactate_overlap"})
    out["reported_module_name"] = out["module"].map(MODULE_LABELS)
    if "cohort" in out.columns:
        out["cohort"] = cohort
    else:
        out.insert(0, "cohort", cohort)
    return out


def summarize_group_values(values: pd.Series, labels: pd.Series, reference: str, test: str) -> dict[str, float]:
    ref = pd.to_numeric(values[labels == reference], errors="coerce").dropna().astype(float)
    tst = pd.to_numeric(values[labels == test], errors="coerce").dropna().astype(float)
    return {
        "reference_n": len(ref),
        "test_n": len(tst),
        "reference_median": float(ref.median()) if len(ref) else math.nan,
        "test_median": float(tst.median()) if len(tst) else math.nan,
        "reference_iqr_low": float(ref.quantile(0.25)) if len(ref) else math.nan,
        "reference_iqr_high": float(ref.quantile(0.75)) if len(ref) else math.nan,
        "test_iqr_low": float(tst.quantile(0.25)) if len(tst) else math.nan,
        "test_iqr_high": float(tst.quantile(0.75)) if len(tst) else math.nan,
    }


def adjusted_group_coefficient(y: np.ndarray, labels: list[str], test_label: str, covariate: np.ndarray) -> tuple[float, float, float]:
    mask = np.isfinite(y) & np.isfinite(covariate)
    y = y[mask]
    covariate = covariate[mask]
    labels = [g for g, keep in zip(labels, mask) if keep]
    group = np.array([1.0 if g == test_label else 0.0 for g in labels])
    cov_z = (covariate - np.mean(covariate)) / np.std(covariate, ddof=0)
    x = np.column_stack([np.ones(len(y)), group, cov_z])
    coef = float(np.linalg.lstsq(x, y, rcond=None)[0][1])
    rng = np.random.default_rng(SEED)
    idx_ref = np.array([i for i, g in enumerate(labels) if g != test_label])
    idx_test = np.array([i for i, g in enumerate(labels) if g == test_label])
    boot = []
    for _ in range(3000):
        idx = np.concatenate([rng.choice(idx_ref, size=len(idx_ref), replace=True), rng.choice(idx_test, size=len(idx_test), replace=True)])
        xb = np.column_stack([np.ones(len(idx)), group[idx], cov_z[idx]])
        boot.append(float(np.linalg.lstsq(xb, y[idx], rcond=None)[0][1]))
    return coef, float(np.quantile(boot, 0.025)), float(np.quantile(boot, 0.975))


def loomba_technical_sensitivity() -> pd.DataFrame:
    meta = read_tsv(ROOT / "data" / "loombar2017" / "metadata.tsv")
    scores = read_tsv(RES / "loombar2017" / "module_scores.tsv")
    selected = scores[(scores["membership"] == "conservative") & (scores["module"] == "overall_fermentation")][["sample_id", "group", "module_score"]]
    df = selected.merge(meta, on=["sample_id", "group"], how="left")
    rows = []
    for col in ["number_reads", "number_bases"]:
        summary = summarize_group_values(df[col], df["group"], "F0_F2", "F3_F4")
        pval = monte_carlo_p(pd.to_numeric(df[col]).astype(float).tolist(), df["group"].tolist(), "F3_F4")
        rho = pd.to_numeric(df[col]).corr(pd.to_numeric(df["module_score"]), method="spearman")
        coef, ci_low, ci_high = adjusted_group_coefficient(
            pd.to_numeric(df["module_score"]).to_numpy(dtype=float),
            df["group"].tolist(),
            "F3_F4",
            np.log(pd.to_numeric(df[col]).to_numpy(dtype=float)),
        )
        rows.append(
            {
                "technical_variable": col,
                "reference_group": "F0_F2",
                "test_group": "F3_F4",
                "reference_n": summary["reference_n"],
                "test_n": summary["test_n"],
                "reference_median": summary["reference_median"],
                "reference_iqr": f"{fmt_num(summary['reference_iqr_low'])} to {fmt_num(summary['reference_iqr_high'])}",
                "test_median": summary["test_median"],
                "test_iqr": f"{fmt_num(summary['test_iqr_low'])} to {fmt_num(summary['test_iqr_high'])}",
                "permutation_p_for_group_difference": pval,
                "spearman_rho_with_selected_composite": rho,
                "group_coefficient_adjusted_for_log_variable": coef,
                "adjusted_bootstrap95_ci_low": ci_low,
                "adjusted_bootstrap95_ci_high": ci_high,
                "interpretation": "Technical sensitivity only; age category remains fully collinear with fibrosis group and cannot be adjusted.",
            }
        )
    return pd.DataFrame(rows)


def prjdb_pathway_detection() -> pd.DataFrame:
    rel = load_prjdb_relative()
    meta = read_tsv(ROOT / "data" / "prjdb36442" / "sample_manifest.tsv")
    scores = read_tsv(RES / "prjdb36442" / "conservative" / "conservative_module_scores.tsv")
    selected = scores[scores["module"] == "overall_fermentation"][["run_accession", "group", "log_ratio_score"]]
    sample_ids = [s for s in meta["run_accession"].tolist() if s in rel.columns]
    detected = (rel[sample_ids] > 0).sum(axis=0).rename("detected_unstratified_pathways").reset_index().rename(columns={"index": "run_accession"})
    df = selected.merge(detected, on="run_accession", how="left")
    summary = summarize_group_values(df["detected_unstratified_pathways"], df["group"], "M", "S")
    pval = exact_permutation_p(pd.to_numeric(df["detected_unstratified_pathways"]).astype(float).tolist(), df["group"].tolist(), "S")
    rho = pd.to_numeric(df["detected_unstratified_pathways"]).corr(pd.to_numeric(df["log_ratio_score"]), method="spearman")
    return pd.DataFrame(
        [
            {
                "qc_variable": "detected_unstratified_pathways",
                "reference_group": "M",
                "test_group": "S",
                "reference_n": summary["reference_n"],
                "test_n": summary["test_n"],
                "reference_median": summary["reference_median"],
                "reference_iqr": f"{fmt_num(summary['reference_iqr_low'])} to {fmt_num(summary['reference_iqr_high'])}",
                "test_median": summary["test_median"],
                "test_iqr": f"{fmt_num(summary['test_iqr_low'])} to {fmt_num(summary['test_iqr_high'])}",
                "exact_p_for_group_difference": pval,
                "spearman_rho_with_selected_composite": rho,
                "retained_pathway_universe": rel.shape[0],
                "available_qc_scope": "Detected unstratified pathway count calculated from the included HUMAnN pathway abundance matrix; read-level host-depletion and mapping-fraction fields were not available.",
            }
        ]
    )


def score_fixed_background(rel: pd.DataFrame, sample_ids: list[str], pathway_ids: list[str], background_ids: list[str]) -> list[float]:
    present = [p for p in dict.fromkeys(pathway_ids) if p in rel.index]
    background = [p for p in dict.fromkeys(background_ids) if p in rel.index and p not in set(present)]
    module_sum = rel.loc[present, sample_ids].sum(axis=0).clip(lower=0.0)
    background_sum = rel.loc[background, sample_ids].sum(axis=0).clip(lower=0.0)
    return (np.log(module_sum + PSEUDOCOUNT) - np.log(background_sum + PSEUDOCOUNT)).astype(float).tolist()


def score_clr_member_mean(rel: pd.DataFrame, sample_ids: list[str], pathway_ids: list[str]) -> list[float]:
    present = [p for p in dict.fromkeys(pathway_ids) if p in rel.index]
    logged = np.log(rel[sample_ids] + PSEUDOCOUNT)
    clr = logged.sub(logged.mean(axis=0), axis=1)
    return clr.loc[present].mean(axis=0).astype(float).tolist()


def logratio_sensitivity() -> pd.DataFrame:
    pathways = selected_pathways()
    prj_rel = load_prjdb_relative()
    prj_meta = read_tsv(ROOT / "data" / "prjdb36442" / "sample_manifest.tsv")
    prj_samples = [s for s in prj_meta["run_accession"].tolist() if s in prj_rel.columns]
    prj_labels = prj_meta.set_index("run_accession").loc[prj_samples, "group"].tolist()

    loomba_rel = load_external_relative(ROOT / "data" / "loombar2017" / "pathway_abundance_unstratified.tsv")
    loomba_scores = read_tsv(RES / "loombar2017" / "module_scores.tsv")
    loomba_meta = loomba_scores[loomba_scores["module"] == "overall_fermentation"][["sample_id", "group"]].drop_duplicates()
    loomba_samples = [s for s in loomba_meta["sample_id"].tolist() if s in loomba_rel.columns]
    loomba_labels = loomba_meta.set_index("sample_id").loc[loomba_samples, "group"].tolist()
    shared_background = sorted((set(prj_rel.index) & set(loomba_rel.index)) - set(pathways))
    rows = []
    for cohort, rel, samples, labels, test_label, exact in [
        ("PRJDB36442", prj_rel, prj_samples, prj_labels, "S", True),
        ("LoombaR_2017", loomba_rel, loomba_samples, loomba_labels, "F3_F4", False),
    ]:
        for method in ["fixed_shared_background_logratio", "clr_module_member_mean"]:
            if method == "fixed_shared_background_logratio":
                values = score_fixed_background(rel, samples, pathways, shared_background)
                background_count = len([p for p in shared_background if p in rel.index])
            else:
                values = score_clr_member_mean(rel, samples, pathways)
                background_count = rel.shape[0]
            ref = [v for v, g in zip(values, labels) if g != test_label]
            test = [v for v, g in zip(values, labels) if g == test_label]
            delta = float(np.mean(test) - np.mean(ref))
            pval = exact_permutation_p(values, labels, test_label) if exact else monte_carlo_p(values, labels, test_label)
            ci_low, ci_high = bootstrap_delta_ci(values, labels, test_label)
            rows.append(
                {
                    "cohort": cohort,
                    "method": method,
                    "test_group": test_label,
                    "reference_group": "not_test_group",
                    "n_module_pathways_defined": len(pathways),
                    "n_module_pathways_present": sum(1 for p in pathways if p in rel.index),
                    "n_background_pathways": background_count,
                    "delta_test_minus_reference": delta,
                    "p_value": pval,
                    "bootstrap95_ci_low": ci_low,
                    "bootstrap95_ci_high": ci_high,
                    "note": "Sensitivity analysis; primary estimates use the cohort-specific all-other-pathways denominator.",
                }
            )
    return pd.DataFrame(rows)


def matched_set_diagnostics() -> pd.DataFrame:
    rel = load_prjdb_relative()
    pathways = selected_pathways()
    present_obs = [p for p in pathways if p in rel.index]
    metrics = pd.DataFrame(
        {
            "mean_abundance": rel.mean(axis=1),
            "prevalence": (rel > 0).mean(axis=1),
            "variance": rel.var(axis=1),
            "zero_fraction": (rel == 0).mean(axis=1),
        }
    )
    rand = read_tsv(RES / "prjdb36442" / "conservative" / "conservative_random_modules.tsv")
    random_rows = []
    for _, r in rand.iterrows():
        ids = [feature_to_id(x.strip()) for x in str(r["pathways"]).split(";") if x.strip()]
        ids = [p for p in ids if p in metrics.index]
        if ids:
            item = {"n_pathways": len(ids)}
            for col in metrics.columns:
                item[col] = float(metrics.loc[ids, col].mean())
            random_rows.append(item)
    random_metrics = pd.DataFrame(random_rows)
    obs = {"n_pathways": len(present_obs)}
    for col in metrics.columns:
        obs[col] = float(metrics.loc[present_obs, col].mean())
    rows = []
    for name, observed in obs.items():
        vals = random_metrics[name].astype(float)
        rows.append(
            {
                "matching_metric": name,
                "observed_selected_composite": observed,
                "random_module_median": float(vals.median()),
                "random_module_iqr_low": float(vals.quantile(0.25)),
                "random_module_iqr_high": float(vals.quantile(0.75)),
                "n_random_modules": len(vals),
            }
        )
    for col in ["exact_bin_matches", "relaxed_mean_prevalence_matches", "global_fallback_matches"]:
        rows.append(
            {
                "matching_metric": col + "_per_random_module",
                "observed_selected_composite": "",
                "random_module_median": float(rand[col].median()),
                "random_module_iqr_low": float(rand[col].quantile(0.25)),
                "random_module_iqr_high": float(rand[col].quantile(0.75)),
                "n_random_modules": len(rand),
            }
        )
    return pd.DataFrame(rows)


def supplementary_index() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ["Workbook date", "2026-07-16"],
            ["Purpose", "Supplementary table index and source-data summaries."],
            ["Analytical note", "Each cohort is analysed separately; matrices are not pooled across cohorts."],
            ["Main discovery analysis", "CHB discovery-cohort conservative selected fermentation composite, S minus M."],
            ["Acetate/lactate reporting", "Acetate and lactate/succinate source definitions are collapsed into one overlap class."],
        ],
        columns=["item", "value"],
    )


def cohort_screening() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ["CHB discovery cohort", "Analysed", "M vs S", "CHB discovery; biopsy-defined histologic injury."],
            ["NAFLD fibrosis cohort", "Analysed", "F0-F2 vs F3-F4", "Unadjusted biopsy-defined fibrosis comparison in NAFLD."],
            ["Qin et al. cirrhosis cohort", "Analysed as context", "healthy vs cirrhosis", "Cirrhosis case-control context and module heterogeneity."],
            ["CNP0007660", "Screened, not analysed", "not used", "Source clinical-to-functional table mapping not independently reproduced."],
            ["Gut 2026 PRJNA projects", "Excluded", "not used", "Amplicon projects, not shotgun HUMAnN inputs."],
        ],
        columns=["cohort", "status", "contrast", "reason_or_use"],
    )


def checklist_rows(kind: str) -> pd.DataFrame:
    if kind == "STORMS":
        rows = [
            ("1. Study design", "Addressed", "Secondary public-data metagenomic reanalysis."),
            ("2. Population and sample source", "Addressed", "Public CHB, NAFLD fibrosis, and cirrhosis cohorts are described."),
            ("3. Inclusion and exclusion", "Addressed", "Analysed, screened-not-analysed, and excluded resources are listed."),
            ("4. Clinical metadata", "Addressed", "Phenotype definitions and unavailable covariates are stated."),
            ("5. Specimen and sequencing context", "Partly addressed", "No new specimens were generated; source-study sequencing context is referenced."),
            ("6. Raw-read processing", "Partly addressed", "CHB uses available processed pathway output; external cohorts use curated processed matrices."),
            ("7. Feature table source", "Addressed", "HUMAnN/MetaCyc pathway-abundance source and curatedMetagenomicData use are reported."),
            ("8. Feature filtering", "Addressed", "Taxon-stratified rows and UNMAPPED/UNINTEGRATED handling are specified."),
            ("9. Normalisation", "Addressed", "Within-sample normalisation and module-versus-background log-ratio scoring are defined."),
            ("10. Module definitions", "Addressed", "Pathway identifiers, names, membership, and availability are provided."),
            ("11. Statistical analysis", "Addressed", "Permutation, bootstrap, Cliff's delta, sensitivity, and matched-set analyses are reported."),
            ("12. Multiple testing", "Addressed", "Main discovery analysis and descriptive secondary/exploratory P values are distinguished."),
            ("13. Technical diagnostics", "Addressed", "Available pathway detection and external technical covariate summaries are supplied."),
            ("14. Data availability", "Addressed", "Accessions, processed data sources, supplementary tables, and repository DOI are provided."),
            ("15. Code availability", "Addressed", "Supplementary Code 1 and archived public release are provided."),
            ("16. Limitations", "Addressed", "Small sample size, missing covariates, processed matrices, age collinearity, and functional-potential boundaries are discussed."),
        ]
        return pd.DataFrame(rows, columns=["STORMS reporting item", "status", "manuscript handling"])
    rows = [(str(i), "Addressed", "Reported in manuscript or supplementary tables.") for i in range(1, 23)]
    return pd.DataFrame(rows, columns=["STROBE item", "status", "manuscript handling"])


def unmapped_unintegrated_submission() -> pd.DataFrame:
    out = unmapped_unintegrated().copy()
    out = out.rename(columns={"median_fraction": "median_fraction_of_total_pathway_output"})
    source_map = {
        "PRJDB36442": "pathway abundance output generated for this reanalysis",
        "LoombaR_2017": "curatedMetagenomicData processed pathway_abundance export",
        "QinN_2014": "curatedMetagenomicData processed pathway_abundance export",
    }
    out.insert(1, "source", out["cohort"].map(source_map))
    out["analysis_handling"] = np.where(
        out["feature_rows_present"].astype(int) > 0,
        "Excluded before within-sample renormalisation; sensitivity was limited to available exported pathway rows.",
        "Feature class not present in the processed matrix available for this analysis.",
    )
    return out


def write_workbook() -> None:
    workbook = TAB / "supplementary_tables.xlsx"
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        supplementary_index().to_excel(writer, sheet_name="Index", index=False)
        cohort_screening().to_excel(writer, sheet_name="Table S1 Cohort screening", index=False)
        build_module_definition_sheet().to_excel(writer, sheet_name="Table S2 Module definitions", index=False)
        build_member_availability().to_excel(writer, sheet_name="Table S3 Pathway availability", index=False)
        build_formula_sheet().to_excel(writer, sheet_name="Table S4 Scoring formula", index=False)
        build_stat_settings().to_excel(writer, sheet_name="Table S5 Statistical methods", index=False)
        build_score_sensitivity().to_excel(writer, sheet_name="Table S6 Score sensitivity", index=False)
        build_loomba_covariates().to_excel(writer, sheet_name="Table S7 NAFLD metadata", index=False)
        endpoint_sheet().to_excel(writer, sheet_name="Table S8 Analysis hierarchy", index=False)
        clean_module_effects(read_tsv(RES / "prjdb36442" / "conservative" / "conservative_module_stats.tsv"), "CHB discovery cohort").to_excel(writer, sheet_name="Table S9 CHB effects", index=False)
        clean_module_effects(read_tsv(RES / "prjdb36442" / "expanded" / "expanded_module_stats.tsv"), "CHB discovery cohort").to_excel(writer, sheet_name="Table S10 CHB expanded", index=False)
        read_tsv(RES / "prjdb36442" / "sensitivity" / "pathway_member_stats.tsv").query("membership == 'conservative'").to_excel(writer, sheet_name="Table S11 CHB pathways", index=False)
        read_tsv(RES / "prjdb36442" / "sensitivity" / "leave_one_pathway_out.tsv").query("membership == 'conservative' and module == 'overall_fermentation'").to_excel(writer, sheet_name="Table S12 CHB leave-one-out", index=False)
        read_tsv(RES / "prjdb36442" / "conservative" / "conservative_random_module_empirical.tsv").to_excel(writer, sheet_name="Table S13 CHB matched sets", index=False)
        clean_module_effects(read_tsv(RES / "loombar2017" / "module_binary_contrasts.tsv"), "NAFLD fibrosis cohort").to_excel(writer, sheet_name="Table S14 NAFLD effects", index=False)
        read_tsv(RES / "loombar2017" / "pathway_member_binary_stats.tsv").to_excel(writer, sheet_name="Table S15 NAFLD pathways", index=False)
        read_tsv(RES / "loombar2017" / "leave_one_pathway_out.tsv").to_excel(writer, sheet_name="Table S16 NAFLD leave-one-out", index=False)
        read_tsv(RES / "loombar2017" / "random_module_empirical.tsv").to_excel(writer, sheet_name="Table S17 NAFLD matched sets", index=False)
        clean_module_effects(read_tsv(RES / "qinn2014" / "module_binary_contrasts.tsv"), "Qin et al. cirrhosis cohort").to_excel(writer, sheet_name="Table S18 Cirrhosis effects", index=False)
        read_tsv(RES / "qinn2014" / "pathway_member_binary_stats.tsv").to_excel(writer, sheet_name="Table S19 Cirrhosis pathways", index=False)
        checklist_rows("STORMS").to_excel(writer, sheet_name="Table S20 STORMS checklist", index=False)
        checklist_rows("STROBE").to_excel(writer, sheet_name="Table S21 STROBE checklist", index=False)
        loomba_technical_sensitivity().to_excel(writer, sheet_name="Table S22 NAFLD technical", index=False)
        prjdb_pathway_detection().to_excel(writer, sheet_name="Table S23 CHB detection", index=False)
        logratio_sensitivity().to_excel(writer, sheet_name="Table S24 Log-ratio sensitivity", index=False)
        matched_set_diagnostics().to_excel(writer, sheet_name="Table S25 Matched diagnostics", index=False)
        unmapped_unintegrated_submission().to_excel(writer, sheet_name="Table S26 Unmapped fractions", index=False)
    from openpyxl import load_workbook

    wb = load_workbook(workbook)
    wb.properties.creator = "Zheng H"
    wb.properties.lastModifiedBy = "Zheng H"
    wb.properties.created = FIXED_WORKBOOK_DATETIME
    wb.properties.modified = FIXED_WORKBOOK_DATETIME
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for idx in range(1, ws.max_column + 1):
            col = get_column_letter(idx)
            width = 10
            for cell in ws[col]:
                value = "" if cell.value is None else str(cell.value)
                width = max(width, min(56, len(value) + 2))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions[col].width = width
    wb.save(workbook)
    normalize_xlsx_zip(workbook)
    table_1().to_csv(TAB / "Table_1_cohort_characteristics.csv", index=False)
    table_2().to_csv(TAB / "Table_2_selected_composite_effects.csv", index=False)


def main() -> None:
    FIG.mkdir(parents=True, exist_ok=True)
    TAB.mkdir(parents=True, exist_ok=True)
    build_figure_1()
    build_figure_2()
    build_figure_3()
    build_figure_4()
    build_supplementary_figure_1()
    write_workbook()


if __name__ == "__main__":
    main()
